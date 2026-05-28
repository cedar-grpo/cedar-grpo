"""
reporting/excel.py
------------------
Build colour-coded Excel comparison workbooks from summary CSVs.
"""

from __future__ import annotations

import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _metric_columns(df: pd.DataFrame) -> list[str]:
    excluded = {"Checkpoint", "Dataset", "Status", "Word Count"}
    cols: list[str] = []
    for c in df.columns:
        if c in excluded:
            continue
        if c.endswith("_analysis") or c.endswith("_examples"):
            continue
        if pd.api.types.is_numeric_dtype(df[c]):
            cols.append(c)
    return cols


def _status_slug(status: str) -> str:
    normalized = "".join(ch.lower() if ch.isalnum() else "_" for ch in status).strip("_")
    return normalized or "status"


def build_excel_workbook(combined: pd.DataFrame, base_dir: str) -> None:
    """
    Create ``checkpoint_comparison.xlsx`` in *base_dir* with one sheet per
    (metric_col × status) combination showing absolute values and delta vs
    checkpoint-0.
    """
    metric_cols = _metric_columns(combined)

    wb = Workbook()
    wb.remove(wb.active)

    for col in metric_cols:
        if "Status" in combined.columns:
            for status in combined["Status"].unique():
                _build_sheet(
                    wb,
                    combined[combined["Status"] == status],
                    col,
                    f"{col[:20]}_{_status_slug(str(status))[:10]}",
                )
        else:
            _build_sheet(wb, combined, col, col[:31])

    if wb.sheetnames:
        path = os.path.join(base_dir, "checkpoint_comparison.xlsx")
        wb.save(path)
        print(f"[OK] Excel → {path}")


def _build_sheet(wb: Workbook, df: pd.DataFrame, metric_col: str, title: str) -> None:
    if metric_col not in df.columns or "Checkpoint" not in df.columns or "Dataset" not in df.columns:
        return
    try:
        pivot = df.pivot_table(index="Checkpoint", columns="Dataset", values=metric_col, aggfunc="mean")
    except Exception:
        return
    if pivot.empty or 0 not in pivot.index:
        return

    baseline = pivot.loc[0]
    diff = pivot.sub(baseline, axis=1)

    ws = wb.create_sheet(title=title[:31])
    ws.cell(1, 1, "Checkpoint").font = Font(bold=True)
    for ci, ds in enumerate(pivot.columns, 2):
        c = ws.cell(1, ci, ds)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    max_pos = float(diff.max().max() or 1)
    max_neg = float(abs(diff.min().min()) or 1)

    for ri, ckpt in enumerate(pivot.index, 2):
        ws.cell(ri, 1, f"ckpt-{ckpt}").font = Font(bold=True)
        for ci, ds in enumerate(pivot.columns, 2):
            val = pivot.loc[ckpt, ds]
            d = diff.loc[ckpt, ds]
            cell = ws.cell(ri, ci)
            cell.alignment = Alignment(horizontal="center")
            if pd.isna(val):
                cell.value = "N/A"
                continue
            cell.value = f"{val:.3f}" if ckpt == 0 else f"{val:.3f} ({d:+.3f})"
            if ckpt != 0 and not pd.isna(d):
                if d > 0:
                    g = int(150 + 105 * min(d / max_pos, 1))
                    cell.fill = PatternFill(start_color=f"00{g:02X}00", end_color=f"00{g:02X}00", fill_type="solid")
                elif d < 0:
                    r = int(150 + 105 * min(abs(d) / max_neg, 1))
                    cell.fill = PatternFill(start_color=f"{r:02X}0000", end_color=f"{r:02X}0000", fill_type="solid")
