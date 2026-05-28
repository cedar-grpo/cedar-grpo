#!/usr/bin/env python3
"""
Aggregate metrics from checkpoint JSONs into a single table.

Directory structure (example):

checkpoints/
  ckpt-001/
    gsm8k/
      all_cases.json
    art/
      all_cases.json
  ckpt-002/
    gsm8k/
      all_cases.json

Each all_cases.json is either:
  {
      "accuracy": 0.87,
      "f1": 0.80,
      ...
  }
or:
  {
      "metrics": {
          "accuracy": 0.87,
          "f1": 0.80,
          ...
      },
      ...
  }

Output:
  - Prints a pretty table to stdout.
  - Writes an Excel file with columns:
        checkpoint, dataset_metric1, dataset_metric2, ...
    and a sheet name controlled by --run.
  - Cells better than the raw model are highlighted in yellow.
  - Two extra summary columns are added at the end:
        better_metrics_than_raw
        better_datasets_than_raw
"""

import os
import json
import argparse
import csv
import numpy as np
from typing import Dict, Any, List, Tuple, Set, Union
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
import re

from evaluate_aime_raw_vs_finetuned import find_best_checkpoint  

from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload  
from google.oauth2.service_account import Credentials
from google.oauth2.credentials import Credentials as UserCredentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from openpyxl import load_workbook
import io 

os.environ.setdefault("OAUTHLIB_INSECURE_TRANSPORT", "1")

Scalar = Union[int, float, str]


def is_scalar(x: Any) -> bool:
    return isinstance(x, (int, float, str))


def load_metrics_from_json(json_path: str) -> Dict[str, Scalar]:
    """Load scalar metrics from all_cases.json.

    Supports:
      - top-level metrics (keys are metric names)
      - or a 'metrics' sub-dict
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if isinstance(data, dict) and "metrics" in data and isinstance(data["metrics"], dict):
        metrics_dict = data["metrics"]
    else:
        metrics_dict = data

    out: Dict[str, Scalar] = {}
    if isinstance(metrics_dict, dict):
        for k, v in metrics_dict.items():
            if is_scalar(v):
                out[k] = v
    return out


def collect_all_rows(root_dir: str, run: str, best_checkpoint: str = None, model_name: str = "qwen2.5-3B") -> Tuple[List[Dict[str, Scalar]], List[str]]:
    """Walk the checkpoints directory and collect rows + column names.

    Returns:
      rows: list of dicts, each representing one checkpoint
      columns: ordered list of column names (including 'checkpoint')
    """
    rows: List[Dict[str, Scalar]] = []
    all_metric_cols: Set[str] = set()

    if not os.path.isdir(root_dir):
        raise FileNotFoundError(f"Root directory not found: {root_dir}")
    
    ckpt_path = os.path.join(root_dir, run, "raw_model")  
    row: Dict[str, Scalar] = {"checkpoint": f"{model_name}"}
    for dataset_name in sorted(os.listdir(ckpt_path)):
        dataset_path = os.path.join(ckpt_path, dataset_name)
        dataset_name = dataset_name.lower()
        if not os.path.isdir(dataset_path):
            continue

        json_path = os.path.join(dataset_path, "raw_results_train_all.json")
        if not os.path.isfile(json_path):
            continue

        try:
            metrics = load_metrics_from_json(json_path)
        except Exception as e:
            print(f"[WARN] Failed to read {json_path}: {e}")
            continue

        f1_flag = False
        possible_col_names = []
        for metric_name, metric_value in metrics.items():
            if "accuracy" in metric_name or "hamming_accuracy" in metric_name:
                col_name = f"{dataset_name}_acc"
            elif "macro_f1" in metric_name or "f1_macro" in metric_name or "f1" in metric_name:
                col_name = f"{dataset_name}_f1"
                f1_flag = True
            elif "precision" in metric_name:
                col_name = f"{dataset_name}_precision"
            elif "recall" in metric_name:
                col_name = f"{dataset_name}_recall"
            elif "exact_match_accuracy" in metric_name:
                col_name = f"{dataset_name}_EM"
            else:
                continue
            
            possible_col_names.append((col_name, metric_value))
        
        for col_name, metric_value in possible_col_names:
            if f1_flag and "_f1" in col_name:    
                if col_name not in row:
                    row[col_name] = round(metric_value, 4)
                    all_metric_cols.add(col_name)
            elif not f1_flag:
                if col_name not in row:
                    row[col_name] = round(metric_value, 4)
                    all_metric_cols.add(col_name)

    rows.append(row)

    root_dir = os.path.join(root_dir, run)
    if not os.path.isdir(root_dir):
        root_dir += "-Evaluation"
        for dir in os.listdir(root_dir):
            if dir.startswith("dt"):
                root_dir = os.path.join(root_dir, dir)
                break
                
    training_step = [int(dir.split("-")[-1]) for dir in os.listdir(root_dir) if "-" in dir]
    for ckpt_name in [f"checkpoint-{str(dir)}" for dir in sorted(training_step)]:
        ckpt_path = os.path.join(root_dir, ckpt_name)
        if not os.path.isdir(ckpt_path) or "checkpoint" not in ckpt_name:
            continue
        
        if best_checkpoint and ckpt_name == best_checkpoint:
            row: Dict[str, Scalar] = {"checkpoint": ckpt_name+"(best)"}
        else:
            row: Dict[str, Scalar] = {"checkpoint": ckpt_name}

        for dataset_name in sorted(os.listdir(ckpt_path)):
            dataset_path = os.path.join(ckpt_path, dataset_name)
            dataset_name = dataset_name.lower()
            if not os.path.isdir(dataset_path):
                continue
            
            json_path = os.path.join(dataset_path, "all_cases.json")
            if not os.path.isfile(json_path):
                json_path = os.path.join(dataset_path, "all_casses.json")
                if not os.path.isfile(json_path):
                    continue
            
            try:
                metrics = load_metrics_from_json(json_path)
            except Exception as e:
                print(f"[WARN] Failed to read {json_path}: {e}")
                continue
            
            f1_flag = False
            possible_col_names = []
            for metric_name, metric_value in metrics.items():
                if "accuracy" in metric_name or "hamming_accuracy" in metric_name:
                    col_name = f"{dataset_name}_acc"
                elif "macro_f1" in metric_name or "f1_macro" in metric_name or "f1" in metric_name:
                    col_name = f"{dataset_name}_f1"
                    f1_flag = True
                elif "precision" in metric_name:
                    col_name = f"{dataset_name}_precision"
                elif "recall" in metric_name:
                    col_name = f"{dataset_name}_recall"
                elif "exact_match_accuracy" in metric_name:
                    col_name = f"{dataset_name}_EM"
                else:
                    continue
                
                possible_col_names.append((col_name, metric_value))
            
            for col_name, metric_value in possible_col_names:
                if f1_flag and "f1" in col_name:    
                    if col_name not in row:
                        row[col_name] = round(metric_value, 4)
                        all_metric_cols.add(col_name)
                elif not f1_flag:
                    if col_name not in row:
                        row[col_name] = round(metric_value, 4)
                        all_metric_cols.add(col_name)

        rows.append(row)

    ordered_cols = ["checkpoint"] + sorted(all_metric_cols)
    return rows, ordered_cols


def format_value(v: Scalar) -> str:
    """Format a scalar value nicely for the ASCII table."""
    if isinstance(v, float):
        return f"{v:.4f}"
    return str(v)


def write_csv(rows: List[Dict[str, Scalar]], columns: List[str], out_path: str) -> None:
    """Write data to a CSV file."""
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            full_row = {col: row.get(col, "") for col in columns}
            writer.writerow(full_row)
    print(f"CSV written to: {out_path}")


def clean_sheet_name(name):
    invalid = ['\\', '/', '*', '?', ':', '[', ']']
    for c in invalid:
        name = name.replace(c, '_')
    # return name[:31]  # Excel also limits sheet names to 31 chars
    return "Sheet1"

def _checkpoint_key(name: str) -> str:
    """Normalize checkpoint name by stripping '(best)' etc."""
    return name.split("(")[0].strip() if name else ""


def append_rows_in_place(
    rows: List[Dict[str, Scalar]],
    columns: List[str],
    out_path: str,
    sheet_name: str,
    best_checkpoint: str,
    model_name: str,
    old_sheet_name: str
) -> None:
    """
    Update an existing Excel sheet in-place:

    - For checkpoints that already exist in the sheet:
        * update all metric cells if values changed
        * recompute better_metrics_than_raw, better_datasets_than_raw
        * update yellow highlighting
    - For new checkpoints:
        * append new rows
    - Preserve column widths and existing sheet (no full rewrite).
    """
    sheet_name = clean_sheet_name(old_sheet_name)

    final_columns = columns + ["better_datasets_than_raw"]
    metric_cols = [c for c in columns if c != "checkpoint"]

    dataset_to_cols: Dict[str, List[str]] = {}
    for col in metric_cols:
        ds = col.split("_", 1)[0]
        dataset_to_cols.setdefault(ds, []).append(col)

    baseline = None
    for r in rows:
        if r.get("checkpoint") == model_name:
            baseline = r
            break

    wb = load_workbook(out_path)

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        for idx, col_name in enumerate(final_columns, start=1):
            ws.cell(row=1, column=idx, value=col_name)

    col_widths = {
        col_letter: dim.width
        for col_letter, dim in ws.column_dimensions.items()
    }

    header_map: Dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            header_map[str(cell.value)] = idx

    max_col = ws.max_column
    for col_name in final_columns:
        if col_name not in header_map:
            max_col += 1
            ws.cell(row=1, column=max_col, value=col_name)
            header_map[col_name] = max_col

    col_index = header_map
    first_data_row = 2

    existing_row_map: Dict[str, int] = {}
    mx_row = 0
    chk_col = col_index["checkpoint"]
    for r_idx in range(first_data_row, ws.max_row + 1):
        val = ws.cell(row=r_idx, column=chk_col).value
        if val:
            ck = _checkpoint_key(str(val))
            if ck not in existing_row_map:
                existing_row_map[ck] = r_idx
                mx_row = max(mx_row, r_idx)

    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    no_fill = PatternFill() 
    best_font = Font(color="008000")
    normal_font = Font(color="000000")

    for row_data in rows:
        ckpt_val = row_data.get("checkpoint")
        if ckpt_val is None:
            continue

        ck_key = _checkpoint_key(str(ckpt_val))

        better_metrics_count = 0
        better_datasets_count = 0
        metric_better_flags: Dict[str, bool] = {}

        if baseline is not None:
            for col in metric_cols:
                v = row_data.get(col)
                b = baseline.get(col)
                improved = False
                try:
                    if v is not None and b is not None:
                        if float(v) > float(b):
                            improved = True
                except Exception:
                    improved = False
                metric_better_flags[col] = improved

            better_metrics_count = sum(metric_better_flags.values())

            for ds, ds_cols in dataset_to_cols.items():
                if any(metric_better_flags.get(c, False) for c in ds_cols):
                    better_datasets_count += 1

        if ck_key in existing_row_map:
            excel_row = existing_row_map[ck_key]
        else:
            excel_row = mx_row + 1
            mx_row += 1
            existing_row_map[ck_key] = excel_row

        for col_name in final_columns:
            idx = col_index[col_name]
            cell = ws.cell(row=excel_row, column=idx)

            if col_name == "better_metrics_than_raw":
                value = better_metrics_count
            elif col_name == "better_datasets_than_raw":
                value = better_datasets_count
            else:
                value = row_data.get(col_name, "")

            cell.value = value

            cell.alignment = Alignment(horizontal='center', vertical='center') 

            if col_name in metric_cols:
                if metric_better_flags.get(col_name, False):
                    cell.fill = green_fill
                else:
                    cell.fill = no_fill

        checkpoint_cell = ws.cell(row=excel_row, column=chk_col)
        if best_checkpoint and _checkpoint_key(best_checkpoint) == ck_key:
            checkpoint_cell.font = best_font
        else:
            checkpoint_cell.font = normal_font

    for col_letter, width in col_widths.items():
        if width is not None:
            ws.column_dimensions[col_letter].width = width

    wb.save(out_path)
    wb.close()


def write_excel(
    rows: List[Dict[str, Scalar]],
    columns: List[str],
    out_path: str,
    sheet_name: str = "Sheet1",
    old_sheet_name: str = None,
    best_checkpoint: str = None,
    model_name: str = "qwen2.5-3B"
) -> None:
    """Write data to an Excel file (one sheet named by sheet_name).

    - If the target sheet already exists, we DO NOT touch its formatting:
      we just append new rows via append_rows_in_place().
    - Otherwise, create the sheet as before.
    - Cells where checkpoint metric > raw_model metric are colored yellow.
    - Two extra columns are appended:
        better_metrics_than_raw
        better_datasets_than_raw
    """
    sheet_name = clean_sheet_name(old_sheet_name)

    if os.path.exists(out_path):
        try:
            wb = load_workbook(out_path)
            if sheet_name in wb.sheetnames:
                wb.close()
                append_rows_in_place(
                    rows=rows,
                    columns=columns,
                    out_path=out_path,
                    sheet_name=sheet_name,
                    best_checkpoint=best_checkpoint,
                    model_name=model_name,
                )
                print(
                    f"Excel updated by appending rows to existing sheet "
                    f"without changing formatting: {out_path} (sheet: {sheet_name})"
                )
                return
            wb.close()
        except Exception as e:
            print(f"[WARN] Could not inspect existing workbook: {e}. Recreating sheet.")

    table = [{col: row.get(col, "") for col in columns} for row in rows]
    df = pd.DataFrame(table, columns=columns)

    metric_cols = [c for c in df.columns if c != "checkpoint"]

    if metric_cols:
        df[metric_cols] = df[metric_cols].apply(pd.to_numeric, errors="coerce")
        df[metric_cols] = np.round(df[metric_cols], 4)

    has_baseline = False
    better_mask = None  

    if "checkpoint" in df.columns and (df["checkpoint"] == f"{model_name}").any() and metric_cols:
        has_baseline = True
        raw_idx = df.index[df["checkpoint"] == f"{model_name}"][0]
        raw_values = df.loc[raw_idx, metric_cols]

        better_mask = df[metric_cols].gt(raw_values)
        better_mask = better_mask.fillna(False)

        # df["better_metrics_than_raw"] = better_mask.sum(axis=1)

        dataset_to_cols = {}
        for col in metric_cols:
            dataset_name = col.split("_", 1)[0]
            dataset_to_cols.setdefault(dataset_name, []).append(col)

        better_datasets_counts: List[int] = []
        for idx, row_bool in better_mask.iterrows():
            count = 0
            for ds, ds_cols in dataset_to_cols.items():
                if row_bool[ds_cols].any():
                    count += 1
            better_datasets_counts.append(count)
        df["better_datasets_than_raw"] = better_datasets_counts
    else:
        # df["better_metrics_than_raw"] = 0
        df["better_datasets_than_raw"] = 0

    final_columns = columns + ["better_datasets_than_raw"]

    mode = "a" if os.path.exists(out_path) else "w"

    if mode == "a":
        with pd.ExcelWriter(out_path, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name, columns=final_columns)

            ws = writer.sheets[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            if has_baseline and better_mask is not None and metric_cols:
                ws = writer.sheets[sheet_name]
                green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

                col_index_map = {col: idx + 1 for idx, col in enumerate(final_columns)}

                for r_idx, row_label in enumerate(df.index):
                    excel_row = r_idx + 2
                    for col in metric_cols:
                        if bool(better_mask.loc[row_label, col]):
                            excel_col = col_index_map[col]
                            ws.cell(row=excel_row, column=excel_col).fill = green_fill
    else:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name, columns=final_columns)

            ws = writer.sheets[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            if has_baseline and better_mask is not None and metric_cols:
                ws = writer.sheets[sheet_name]
                green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                col_index_map = {col: idx + 1 for idx, col in enumerate(final_columns)}

                for r_idx, row_label in enumerate(df.index):
                    excel_row = r_idx + 2
                    for col in metric_cols:
                        if bool(better_mask.loc[row_label, col]):
                            excel_col = col_index_map[col]
                            ws.cell(row=excel_row, column=excel_col).fill = green_fill
            
            if best_checkpoint:
                ws = writer.sheets[sheet_name]
                checkpoint_col = final_columns.index("checkpoint") + 1
                for r_idx, row_label in enumerate(df.index):
                    if df.loc[row_label, "checkpoint"] == best_checkpoint:
                        excel_row = r_idx + 2
                        cell = ws.cell(row=excel_row, column=checkpoint_col)
                        cell.font = Font(color="008000")

    print(f"Excel written to: {out_path} (sheet: {sheet_name})")

GDRIVE_TOKEN_PATH = os.path.expanduser("~/.config/nima_drive_token.json")
GDRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.file"]


def get_oauth_credentials(oauth_client_json: str) -> UserCredentials:
    """
    Get (or create) OAuth user credentials for Google Drive.

    - Reuses a saved token if it exists.
    - Otherwise:
        * Builds an auth URL with a proper redirect_uri
        * Asks you to open it and grant access
        * You paste the FULL redirect URL from the browser
        * Exchanges that for tokens and saves them
    """
    creds = None

    if os.path.exists(GDRIVE_TOKEN_PATH):
        creds = UserCredentials.from_authorized_user_file(
            GDRIVE_TOKEN_PATH, GDRIVE_SCOPES
        )

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            with open(oauth_client_json, "r", encoding="utf-8") as f:
                cfg = json.load(f)

            flow = InstalledAppFlow.from_client_secrets_file(
                oauth_client_json,
                scopes=GDRIVE_SCOPES,
            )

            redirect_uri = None
            if "installed" in cfg and "redirect_uris" in cfg["installed"]:
                redirect_uri = cfg["installed"]["redirect_uris"][0]
            elif "web" in cfg and "redirect_uris" in cfg["web"]:
                redirect_uri = cfg["web"]["redirect_uris"][0]

            if redirect_uri is None:
                redirect_uri = "http://localhost"

            flow.redirect_uri = redirect_uri

            auth_url, _ = flow.authorization_url(
                access_type="offline",
                include_granted_scopes="true",
                prompt="consent",
            )

            print("\n🔐 Google Drive authorization required.")
            print("1) Open this URL in your browser:\n")
            print(auth_url)
            print(
                "\n2) Approve access. Google will then redirect you to something "
                "like:\n"
                "   http://localhost/?code=...&scope=...\n"
                "   (The page may show a connection error; that's fine.)"
            )
            print(
                "\n3) Copy the FULL redirect URL from your browser's address bar "
                "and paste it here."
            )
            redirect_response = input("\nRedirect URL: ").strip()

            flow.fetch_token(authorization_response=redirect_response)
            creds = flow.credentials

        os.makedirs(os.path.dirname(GDRIVE_TOKEN_PATH), exist_ok=True)
        with open(GDRIVE_TOKEN_PATH, "w", encoding="utf-8") as token_file:
            token_file.write(creds.to_json())

    return creds

def download_from_gdrive_if_exists(
    local_path: str,
    folder_id: str,
    oauth_client_json: str,
    sheet_name: str
) -> bool:
    """
    If a file with the given name exists in the Drive folder, download it
    to local_path and return True. Otherwise, do nothing and return False.
    """
    creds = get_oauth_credentials(oauth_client_json)
    service = build("drive", "v3", credentials=creds)

    # file_name = os.path.basename(local_path)
    file_name = sheet_name
    query = (
        f"name = '{file_name}' and "
        f"'{folder_id}' in parents and "
        f"trashed = false"
    )

    existing = (
        service.files()
        .list(q=query, spaces="drive", fields="files(id, name)", pageSize=1)
        .execute()
    )
    files = existing.get("files", [])
    if not files:
        return False 

    file_id = files[0]["id"]
    request = service.files().get_media(fileId=file_id)

    fh = io.FileIO(local_path, "wb")
    downloader = MediaIoBaseDownload(fh, request)

    done = False
    while not done:
        status, done = downloader.next_chunk()

    return True


def upload_to_gdrive(
    local_path: str,
    folder_id: str,
    oauth_client_json: str,
    mime_type: str = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    sheet_name: str = "Sheet1",
) -> str:
    """
    Upload a local file to a Google Drive folder using OAuth user credentials.

    If a file with the same name already exists in that folder:
      -> update its contents (overwrite) instead of creating a new file.
    """
    if not os.path.isfile(local_path):
        raise FileNotFoundError(f"File not found for upload: {local_path}")

    creds = get_oauth_credentials(oauth_client_json)
    service = build("drive", "v3", credentials=creds)

    # file_name = os.path.basename(local_path)
    file_name = sheet_name

    query = (
        f"name = '{file_name}' and "
        f"'{folder_id}' in parents and "
        f"trashed = false"
    )

    existing = (
        service.files()
        .list(q=query, spaces="drive", fields="files(id, name)", pageSize=1)
        .execute()
    )
    files = existing.get("files", [])

    media = MediaFileUpload(local_path, mimetype=mime_type, resumable=True)

    if files:
        file_id = files[0]["id"]
        print(f"Updating existing file on Drive: {file_name} (id={file_id})")
        uploaded = (
            service.files()
            .update(fileId=file_id, media_body=media, fields="id")
            .execute()
        )
    else:
        file_metadata = {
            "name": file_name,
            "parents": [folder_id],
        }
        print(f"Creating new file on Drive: {file_name}")
        uploaded = (
            service.files()
            .create(body=file_metadata, media_body=media, fields="id")
            .execute()
        )

    file_id = uploaded.get("id")
    print(f"Uploaded to Google Drive. File ID: {file_id}")
    return file_id


def write_excel_to_gdrive(
    rows: List[Dict[str, Scalar]],
    columns: List[str],
    out_path: str,
    sheet_name: str,
    old_sheet_name: str,
    best_checkpoint: str,
    model_name: str,
    gdrive_folder_id: str,
    gdrive_service_account_json: str,
) -> None:
    """
    - If the file does not exist on Drive:
        * create a new Excel file locally with all rows (write_excel)
        * upload it
    - If it exists:
        * download latest version (including manual Google formatting)
        * append only new rows to the existing sheet
        * upload updated file back to Drive
    """
    exists_remote = download_from_gdrive_if_exists(
        local_path=out_path,
        folder_id=gdrive_folder_id,
        oauth_client_json=gdrive_service_account_json,
        sheet_name=sheet_name
    )

    if not exists_remote and not os.path.exists(out_path):
        write_excel(
            rows=rows,
            columns=columns,
            out_path=out_path,
            sheet_name=sheet_name,
            old_sheet_name=old_sheet_name,
            best_checkpoint=best_checkpoint,
            model_name=model_name,
        )
    else:
        append_rows_in_place(
            rows=rows,
            columns=columns,
            out_path=out_path,
            sheet_name=sheet_name,
            old_sheet_name=old_sheet_name,
            best_checkpoint=best_checkpoint,
            model_name=model_name,
        )

    upload_to_gdrive(
        local_path=out_path,
        folder_id=gdrive_folder_id,
        oauth_client_json=gdrive_service_account_json,
        sheet_name=sheet_name,
    )



def main():
    parser = argparse.ArgumentParser(
        description="Create a metrics table from checkpoint JSON files."
    )
    parser.add_argument(
        "--root",
        type=str,
        default="./GRPO/Evaluation/",
        help="Root directory containing checkpoints (default: checkpoints)",
    )
    parser.add_argument(
        "--out_csv",
        type=str,
        default="./GRPO/Evaluation//metrics_summary.xlsx",
        help="Output Excel file path (default: metrics_summary.xlsx)",
    )
    parser.add_argument(
        "--run",
        type=str,
        default="dt11.18.17:40_e20_unsloth_Qwen2.5_3B_Instruct_unsloth_bnb_4bit_bnb_4bit_lr1e-05_t0.7_ε0.2_r64_b16",
        help="Name of the Excel sheet (subsheet) to write results into.",
    )
    parser.add_argument(
        "--best_checkpoint",
        type=str,
        default=None,
        help="Name of checkpoint whose name in the first column will be colored green.",
    )
    parser.add_argument(
        "--base_model_name",
        type=str,
        default="qwen2.5-3B",
        help="Name of the base model we trained on",
    )
    parser.add_argument(
        "--base_result_dir",
        type=str,
        default="~/users/Nima/AbductiveReasoning/GRPO/results",
        help="Directory of the base model we trained on",
    )
    parser.add_argument(
        "--train_data",
        type=str,
        default="UniADILR",
        help="Name of the training data that the model was trained on.",
    )

    args = parser.parse_args()

    if args.best_checkpoint is None:
        BASE_RESULTS_DIR = args.base_result_dir
        TRAINING_DIR=f"{BASE_RESULTS_DIR}/Training_{args.run}"
        FINAL_DIR=f"{BASE_RESULTS_DIR}/{args.run}"
        if os.path.isdir(TRAINING_DIR):
            TRAINING_BASE = TRAINING_DIR
        elif os.path.isdir(FINAL_DIR):
            TRAINING_BASE = FINAL_DIR
        else:
            print(f"ERROR: Could not find checkpoint directory.")
            print(f"Tried:")
            print(f"  {TRAINING_DIR}")
            print(f"  {FINAL_DIR}")
            return 

        best_path, _ = find_best_checkpoint(TRAINING_BASE)
        best_path, _ = find_best_checkpoint(TRAINING_BASE)
        args.best_checkpoint = os.path.basename(best_path) if best_path else None
        # args.best_checkpoint = "checkpoint-4096"

    rows, columns = collect_all_rows(args.root, args.run, args.best_checkpoint, args.base_model_name)
    # write_csv(rows, columns, args.out_csv)
    old_sheet_name = args.run

    match = re.search(r"(_e\d+_)", args.run)
    if match:
        start, end = match.span()
        # Insert train_data before the epoch part
        sheet_name = args.run[:start] + "_" + args.train_data + args.run[start:]
    else:
        # Fallback if pattern not found
        sheet_name = args.run + "_" + args.train_data
    # Google Drive service account JSON path - should be set via environment variable or config
    gdrive_json = os.environ.get('GDRIVE_SERVICE_ACCOUNT_JSON', 
                                  os.path.expanduser("~/client_secret_709163142430-45tbm173bvr506elk6mvf1093ecatcmg.apps.googleusercontent.com.json"))
    write_excel_to_gdrive(rows, columns, args.out_csv, sheet_name=sheet_name, old_sheet_name=old_sheet_name, best_checkpoint=args.best_checkpoint, model_name=args.base_model_name, gdrive_folder_id="1UVSy7yB2pvj8GSa9ns89JAujxzxkLEC-", gdrive_service_account_json=gdrive_json) 


if __name__ == "__main__":
    main()